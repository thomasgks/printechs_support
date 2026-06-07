from __future__ import annotations

from io import BytesIO
import re
from html import escape as html_escape
from urllib.parse import quote

import frappe
from frappe import _
from frappe.utils import cint, date_diff, format_datetime, get_url, getdate, now_datetime, strip_html, today
from frappe.utils.xlsxutils import make_xlsx
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PENDING_TICKET_STATUSES = (
	"Open",
	"Assigned",
	"In Progress",
	"Waiting for Customer",
	"Waiting for Technician",
	"Reopened",
)

REPORT_COLUMNS = (
	("ticket", _("Ticket")),
	("subject", _("Subject")),
	("customer_name", _("Customer Name")),
	("status", _("Status")),
	("priority", _("Priority")),
	("ticket_type", _("Ticket Type")),
	("opening_date", _("Opening Date")),
	("due_date", _("Due Date")),
	("first_response_on", _("First Response Date")),
	("resolution_due", _("Resolution Date")),
	("resolved_on", _("Resolved On")),
	("closed_on", _("Closed On")),
	("delay_reason", _("Delay Reason")),
	("delay_remarks", _("Delay Remarks")),
)

EXCEL_ONLY_COLUMNS = (
	("last_communication_on", _("Last Communication Date")),
	("last_communication_by", _("Last Communication By")),
	("last_communication_from_type", _("Last Communication From Type")),
	("last_communication", _("Last Communication")),
	("last_reply_on", _("Reply Date")),
	("last_reply_by", _("Reply By")),
	("last_reply_from_type", _("Reply From Type")),
	("reply_delay_hours", _("Reply Delay (Hours)")),
	("reply_delay_days", _("Reply Delay (Days)")),
	("last_reply", _("Reply")),
)

STATUS_FILL_COLORS = {
	"Open": "DBEAFE",
	"Assigned": "EDE9FE",
	"In Progress": "FEF3C7",
	"Waiting for Customer": "FFEDD5",
	"Waiting for Technician": "FCE7F3",
	"Reopened": "FED7AA",
}

ACTOR_TYPE_FILL_COLORS = {
	"Customer": "DCFCE7",
	"Technician": "DBEAFE",
	"Manager": "EDE9FE",
	"System": "E5E7EB",
}


def send_pending_ticket_sla_report_if_due(*, force: bool = False) -> None:
	"""Send the pending-ticket SLA report using the configured day interval."""
	settings = frappe.get_single("Printechs Support Settings")
	last_sent = settings.get("last_pending_ticket_report_sent_on")
	interval_days = max(cint(settings.get("pending_ticket_report_interval_days")) or 2, 1)
	if not force and last_sent and date_diff(getdate(today()), getdate(last_sent)) < interval_days:
		return

	sent_count = 0
	for team in _get_active_support_teams():
		recipients = _get_team_recipients(team, settings)
		if not recipients:
			continue
		rows = get_pending_ticket_report_rows(team=team.name)
		if not rows:
			continue
		add_last_communication_context(rows)
		_send_report(recipients, rows, settings=settings, team_label=team.team_name or team.name)
		sent_count += 1

	fallback_recipients = _split_emails(settings.get("pending_ticket_report_recipients"))
	if fallback_recipients:
		unassigned_rows = get_pending_ticket_report_rows(unassigned=True)
		if unassigned_rows:
			add_last_communication_context(unassigned_rows)
			_send_report(
				fallback_recipients,
				unassigned_rows,
				settings=settings,
				team_label=_("Unassigned Tickets"),
			)
			sent_count += 1

	if sent_count:
		settings.db_set("last_pending_ticket_report_sent_on", now_datetime(), update_modified=False)


@frappe.whitelist()
def send_pending_ticket_sla_report_now() -> dict:
	"""Manual trigger for administrators to test the pending-ticket report email."""
	if not frappe.has_permission("Printechs Support Settings", "write"):
		frappe.throw(_("Not permitted"), frappe.PermissionError)
	send_pending_ticket_sla_report_if_due(force=True)
	return {"ok": True}


def get_pending_ticket_report_rows(team: str | None = None, unassigned: bool = False) -> list[dict]:
	conditions = [
		"st.docstatus < 2",
		"st.status IN %(statuses)s",
	]
	params: dict = {"statuses": PENDING_TICKET_STATUSES}
	if team:
		conditions.append("st.team = %(team)s")
		params["team"] = team
	elif unassigned:
		conditions.append("IFNULL(st.team, '') = ''")

	return frappe.db.sql(
		f"""
		SELECT
			st.name AS ticket,
			st.subject AS subject,
			COALESCE(NULLIF(st.customer_name, ''), st.customer) AS customer_name,
			st.status AS status,
			st.priority AS priority,
			st.ticket_type AS ticket_type,
			st.opening_date AS opening_date,
			st.due_date AS due_date,
			st.first_response_on AS first_response_on,
			st.resolution_due AS resolution_due,
			st.resolved_on AS resolved_on,
			st.closed_on AS closed_on,
			IFNULL(dr.reason_name, st.delay_reason) AS delay_reason,
			st.delay_remarks AS delay_remarks
		FROM `tabSupport Ticket` st
		LEFT JOIN `tabDelay Reason` dr ON dr.name = st.delay_reason
		WHERE {" AND ".join(conditions)}
		ORDER BY
			st.due_date IS NULL ASC,
			st.due_date ASC,
			FIELD(st.priority, 'Critical', 'High', 'Medium', 'Low') ASC,
			st.opening_date DESC
		""",
		params,
		as_dict=True,
	)


def add_last_communication_context(rows: list[dict]) -> None:
	ticket_names = [row.get("ticket") for row in rows if row.get("ticket")]
	if not ticket_names:
		return

	comment_rows = frappe.get_all(
		"Support Ticket Comment",
		filters={
			"parent": ["in", ticket_names],
			"parenttype": "Support Ticket",
			"parentfield": "comments",
			"comment_type": ["!=", "System Update"],
		},
		fields=["name", "parent", "comment_by", "comment_on", "content", "in_reply_to"],
		order_by="parent asc, comment_on asc, creation asc",
		limit_page_length=0,
	)
	by_ticket: dict[str, list[dict]] = {}
	by_name: dict[str, dict] = {}
	children_by_parent: dict[str, list[dict]] = {}
	for comment in comment_rows:
		parent = comment.get("parent")
		if parent:
			by_ticket.setdefault(parent, []).append(comment)
		name = comment.get("name")
		if name:
			by_name[name] = comment
		reply_to = (comment.get("in_reply_to") or "").strip()
		if reply_to:
			children_by_parent.setdefault(reply_to, []).append(comment)

	for row in rows:
		comments = by_ticket.get(row.get("ticket") or "", [])
		if not comments:
			_set_empty_communication_context(row)
			continue

		latest = comments[-1]
		reply_to = (latest.get("in_reply_to") or "").strip()
		if reply_to and reply_to in by_name:
			last_message = by_name[reply_to]
			last_reply = latest
		else:
			last_message = latest
			replies = children_by_parent.get(latest.get("name") or "", [])
			last_reply = replies[-1] if replies else None

		row["last_communication_on"] = last_message.get("comment_on")
		row["last_communication_by"] = _comment_author_label(last_message.get("comment_by"))
		row["last_communication_from_type"] = _comment_actor_type(last_message.get("comment_by"))
		row["last_communication"] = _comment_text_preview(last_message.get("content"))
		if last_reply:
			row["last_reply_on"] = last_reply.get("comment_on")
			row["last_reply_by"] = _comment_author_label(last_reply.get("comment_by"))
			row["last_reply_from_type"] = _comment_actor_type(last_reply.get("comment_by"))
			delay_hours = _datetime_diff_hours(last_message.get("comment_on"), last_reply.get("comment_on"))
			row["reply_delay_hours"] = delay_hours
			row["reply_delay_days"] = round(delay_hours / 24, 2) if delay_hours is not None else ""
			row["last_reply"] = _comment_text_preview(last_reply.get("content"))
		else:
			row["last_reply_on"] = None
			row["last_reply_by"] = ""
			row["last_reply_from_type"] = ""
			row["reply_delay_hours"] = ""
			row["reply_delay_days"] = ""
			row["last_reply"] = ""


def build_pending_ticket_report_html(rows: list[dict], *, settings=None, team_label: str | None = None) -> str:
	count = len(rows)
	generated_on = format_datetime(now_datetime())
	period_note = _("Pending tickets only. Hold, Resolved, Closed, and Cancelled tickets are excluded.")
	custom_message = ""
	if settings:
		custom_message = _format_setting_template(settings.get("pending_ticket_report_message"), rows, team_label)
	custom_message_html = _plain_text_to_html(custom_message) if custom_message else _plain_text_to_html(
		_(
			"Dear Support Team,\n\nPlease find below the pending ticket SLA and delay report. "
			"The Excel attachment contains the same list for filtering and follow-up."
		)
	)

	if not rows:
		body = f"""
<p style="margin:16px 0 0;color:#475569;">{html_escape(_("There are no pending support tickets."))}</p>
"""
	else:
		body = _build_table_html(rows)

	return f"""<div style="font-family:Arial,Helvetica,sans-serif;color:#0f172a;font-size:13px;line-height:1.45;">
<div style="max-width:1180px;">
<h2 style="margin:0 0 6px;font-size:20px;color:#0f172a;">{html_escape(_("Pending Ticket SLA and Delay Report"))}</h2>
<p style="margin:0 0 12px;color:#475569;">
{html_escape(_("Generated on"))}: {html_escape(generated_on)}<br>
{html_escape(_("Support Team"))}: <strong>{html_escape(team_label or _("All Teams"))}</strong><br>
{html_escape(_("Pending tickets"))}: <strong>{count}</strong><br>
{html_escape(period_note)}
</p>
<div style="margin:0 0 14px;color:#334155;background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:12px;">
{custom_message_html}
</div>
{body}
<p style="margin:14px 0 0;color:#64748b;font-size:12px;">
{html_escape(_("The attached Excel file contains the same ticket list for filtering and follow-up."))}
</p>
</div>
</div>"""


def build_pending_ticket_report_xlsx(rows: list[dict]) -> dict:
	columns = REPORT_COLUMNS + EXCEL_ONLY_COLUMNS
	data = [[label for _, label in columns]]
	for row in rows:
		data.append([_excel_value(row.get(field)) for field, _label in columns])
	xlsx = make_xlsx(
		data,
		"Pending Tickets",
		column_widths=[18, 42, 28, 18, 12, 22, 20, 20, 20, 20, 20, 20, 24, 44, 22, 24, 22, 60, 22, 24, 18, 18, 18, 60],
	)
	return {
		"fname": f"pending-ticket-sla-report-{today()}.xlsx",
		"fcontent": _apply_status_cell_fills(xlsx, columns),
	}


def _build_table_html(rows: list[dict]) -> str:
	headers = "".join(
		f'<th style="{_th_style()}">{html_escape(str(label))}</th>' for _field, label in REPORT_COLUMNS
	)
	body_rows = []
	for row in rows:
		cells = []
		for field, _label in REPORT_COLUMNS:
			value = _html_value(row.get(field))
			if field == "ticket" and value != "—":
				url = _portal_ticket_url(value)
				value = f'<a href="{html_escape(url)}" style="color:#2563eb;text-decoration:none;font-weight:600;">{html_escape(value)}</a>'
			else:
				value = html_escape(value)
			cells.append(f'<td style="{_td_style()}">{value}</td>')
		body_rows.append(f"<tr>{''.join(cells)}</tr>")

	return f"""<div style="overflow-x:auto;">
<table style="border-collapse:collapse;width:100%;min-width:1100px;border:1px solid #e2e8f0;">
<thead><tr>{headers}</tr></thead>
<tbody>{''.join(body_rows)}</tbody>
</table>
</div>"""


def _send_report(recipients: list[str], rows: list[dict], *, settings, team_label: str) -> None:
	message = build_pending_ticket_report_html(rows, settings=settings, team_label=team_label)
	attachment = build_pending_ticket_report_xlsx(rows)
	subject = _format_setting_template(
		settings.get("pending_ticket_report_subject") or _("Pending Ticket SLA Report - {team} - {date}"),
		rows,
		team_label,
	)

	frappe.sendmail(
		recipients=recipients,
		subject=subject,
		message=message,
		attachments=[attachment],
		delayed=False,
	)


def _portal_ticket_url(ticket_name: str) -> str:
	base = get_url().rstrip("/")
	name = (ticket_name or "").strip()
	if not name:
		return f"{base}/support-portal"
	return f"{base}/support-portal/tickets/{quote(name)}"


def _get_active_support_teams() -> list[dict]:
	return frappe.get_all(
		"Support Team",
		filters={"is_active": 1},
		fields=["name", "team_name", "default_email", "team_lead_email"],
		order_by="team_name asc, name asc",
	)


def _get_team_recipients(team, settings) -> list[str]:
	recipients = _collect_support_team_emails(team.name)
	if recipients:
		return recipients
	return _split_emails(settings.get("pending_ticket_report_recipients"))


def _format_setting_template(template: str | None, rows: list[dict], team_label: str | None = None) -> str:
	template = (template or "").strip()
	if not template:
		return ""
	try:
		return template.format(date=today(), count=len(rows), team=team_label or _("All Teams"))
	except Exception:
		return template


def _plain_text_to_html(text: str) -> str:
	lines = html_escape(text or "").splitlines()
	if not lines:
		return ""
	return "<br>".join(lines)


def _split_emails(raw: str | None) -> list[str]:
	out: list[str] = []
	for item in re.split(r"[\s,;]+", raw or ""):
		email = _normalize_email(item)
		if email and email not in out:
			out.append(email)
	return out


def _collect_support_team_emails(team_name: str) -> list[str]:
	out: list[str] = []

	def add(email: str | None) -> None:
		email = _normalize_email(email)
		if email and email not in out:
			out.append(email)

	team = frappe.db.get_value(
		"Support Team",
		team_name,
		["default_email", "team_lead_email"],
		as_dict=True,
	)
	if team:
		add(team.default_email)
		add(team.team_lead_email)
	for user in frappe.get_all(
		"Support Team Member",
		filters={"parent": team_name, "parenttype": "Support Team"},
		pluck="user",
	):
		add(frappe.db.get_value("User", user, "email"))

	return out


def _normalize_email(email: str | None) -> str | None:
	email = (email or "").strip()
	if not email:
		return None
	try:
		from frappe.utils import validate_email_address

		validate_email_address(email, True)
		return email.lower()
	except Exception:
		return None


def _html_value(value) -> str:
	if value is None or value == "":
		return "—"
	if hasattr(value, "strftime"):
		return format_datetime(value)
	return str(value)


def _excel_value(value):
	if value is None:
		return ""
	return value


def _apply_status_cell_fills(xlsx_file, columns: tuple[tuple[str, str], ...]) -> bytes:
	status_col_idx = next((idx + 1 for idx, (field, _label) in enumerate(columns) if field == "status"), None)

	xlsx_file.seek(0)
	wb = load_workbook(xlsx_file)
	ws = wb.active
	_style_pending_ticket_sheet(ws, columns)
	for row_idx in range(2, ws.max_row + 1):
		if status_col_idx:
			cell = ws.cell(row=row_idx, column=status_col_idx)
			color = STATUS_FILL_COLORS.get(str(cell.value or "").strip())
			if color:
				cell.fill = PatternFill(fill_type="solid", fgColor=color)

	out = BytesIO()
	wb.save(out)
	return out.getvalue()


def _style_pending_ticket_sheet(ws, columns: tuple[tuple[str, str], ...]) -> None:
	header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
	header_font = Font(color="FFFFFF", bold=True)
	border = Border(
		left=Side(style="thin", color="CBD5E1"),
		right=Side(style="thin", color="CBD5E1"),
		top=Side(style="thin", color="CBD5E1"),
		bottom=Side(style="thin", color="CBD5E1"),
	)
	wrap_fields = {"subject", "delay_remarks", "last_communication", "last_reply"}
	wrap_columns = {idx + 1 for idx, (field, _label) in enumerate(columns) if field in wrap_fields}
	actor_type_columns = {
		idx + 1
		for idx, (field, _label) in enumerate(columns)
		if field in {"last_communication_from_type", "last_reply_from_type"}
	}

	ws.freeze_panes = "A2"
	ws.auto_filter.ref = ws.dimensions
	ws.row_dimensions[1].height = 24

	for cell in ws[1]:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
		cell.border = border

	for row in ws.iter_rows(min_row=2):
		for cell in row:
			if cell.column in actor_type_columns:
				color = ACTOR_TYPE_FILL_COLORS.get(str(cell.value or "").strip())
				if color:
					cell.fill = PatternFill(fill_type="solid", fgColor=color)
			cell.alignment = Alignment(
				vertical="top",
				wrap_text=cell.column in wrap_columns,
			)
			cell.border = border


def _set_empty_communication_context(row: dict) -> None:
	for field, _label in EXCEL_ONLY_COLUMNS:
		row[field] = ""


def _comment_author_label(user: str | None) -> str:
	user = (user or "").strip()
	if not user:
		return ""
	full_name = frappe.db.get_value("User", user, "full_name")
	return (full_name or user).strip()


def _comment_actor_type(user: str | None) -> str:
	user = (user or "").strip()
	if not user:
		return ""
	try:
		from printechs_support.printechs_support_system.api.ticket_workflow import classify_actor_role_type

		actor_type = classify_actor_role_type(user)
		return "Technician" if actor_type == "Manager" else actor_type
	except Exception:
		return "Technician"


def _datetime_diff_hours(start, end) -> float | None:
	if not start or not end:
		return None
	try:
		from frappe.utils import get_datetime

		start_dt = get_datetime(start)
		end_dt = get_datetime(end)
		return round((end_dt - start_dt).total_seconds() / 3600, 2)
	except Exception:
		return None


def _comment_text_preview(content_html: str | None, max_chars: int = 500) -> str:
	text = strip_html(content_html or "")
	text = re.sub(r"\s+", " ", text).strip()
	if not text:
		return ""
	if len(text) > max_chars:
		return text[: max_chars - 1].rstrip() + "..."
	return text


def _th_style() -> str:
	return "background:#f1f5f9;color:#334155;padding:8px;border:1px solid #e2e8f0;text-align:left;font-size:12px;"


def _td_style() -> str:
	return "padding:7px 8px;border:1px solid #e2e8f0;vertical-align:top;color:#0f172a;"
